import io, streamlit as st
from datetime import datetime
from openpyxl import load_workbook
from docx import Document
from report_helper import (
    fill_deal_summary, fill_shareholding, fill_restrict,
    fill_basic_info, fill_delist, fill_stress_test, fill_company_info,
    fill_solvency, fill_turnover, fill_cashflow,
    fill_comparison, fill_industry, fill_financial, fill_legacy,
)

st.set_page_config(page_title='尽调报告自动填表', page_icon='📊', layout='centered')
st.title('尽调报告自动填表工具')

col1, col2 = st.columns(2)
with col1:
    excel_file = st.file_uploader('上传 Excel（.xlsx）', type=['xlsx'])
with col2:
    word_file = st.file_uploader('上传 Word（.docx）', type=['docx'])

if excel_file and word_file:
    if st.button('开始生成', type='primary', use_container_width=True):
        with st.status('处理中...', expanded=True) as status:

            st.write('读取 Excel...')
            excel_bytes = excel_file.read()
            wb_val = load_workbook(io.BytesIO(excel_bytes), data_only=True)
            wb_fmt = load_workbook(io.BytesIO(excel_bytes), data_only=False)

            st.write('读取 Word...')
            doc = Document(io.BytesIO(word_file.read()))

            st.write('申请项目要素...')
            fill_deal_summary(doc, wb_val, wb_fmt)

            st.write('持股概要...')
            fill_shareholding(doc, wb_val, wb_fmt)

            st.write('减持受限...')
            fill_restrict(doc, wb_val, wb_fmt)

            st.write('标的证券基本面...')
            fill_basic_info(doc, wb_val, wb_fmt)

            st.write('退市指标排查...')
            fill_delist(doc, wb_val, wb_fmt)

            st.write('压力测试...')
            fill_stress_test(doc, wb_val, wb_fmt)

            st.write('上市公司概况...')
            fill_company_info(doc, wb_val)

            st.write('偿债能力分析...')
            fill_solvency(doc, wb_val, wb_fmt)

            st.write('资产周转率分析...')
            fill_turnover(doc, wb_val, wb_fmt)

            st.write('现金流分析...')
            fill_cashflow(doc, wb_val, wb_fmt)

            st.write('对标同行业比较...')
            fill_comparison(doc, wb_val)

            st.write('行业估值比较...')
            fill_industry(doc, wb_val)

            st.write('财务分析...')
            fill_financial(doc, wb_val)

            st.write('基础模块...')
            fill_legacy(doc, wb_val, wb_fmt)

            output = io.BytesIO()
            doc.save(output)
            output.seek(0)

            status.update(label='完成！', state='complete')

        ts = datetime.now().strftime('%Y%m%d_%H%M')
        st.download_button(
            label='下载报告',
            data=output,
            file_name=f'合并_尽调报告{ts}.docx',
            mime='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            type='primary',
            use_container_width=True,
        )
else:
    st.info('请上传 Excel 和 Word 文件后点击生成')
